import json
from multiprocessing import Pool

import openpyxl
from osgeo import gdal
import os

from tqdm import tqdm

from UtilitiesForProcessingImage.BasicUtility.UtilityFunction import workdir_filelist


def para_cal(input_file_path, output_file_dir, ref_dictionary):
    print(f"---PID: {os.getpid()}---\n---Dealing with {input_file_path}---")
    # read ds
    in_ds = gdal.Open(input_file_path, gdal.GA_ReadOnly)
    lucc_array = in_ds.GetRasterBand(1).ReadAsArray()
    ds_trans = in_ds.GetGeoTransform()
    ds_proj = in_ds.GetProjection()
    ds_width = in_ds.RasterXSize
    ds_height = in_ds.RasterYSize
    ds_band_num = in_ds.RasterCount
    del in_ds
    # create empty SLA raster
    sla_path = os.path.join(output_file_dir, "SLA_" + os.path.basename(input_file_path)[9:13] + ".tif")
    driver = gdal.GetDriverByName('GTiff')
    sla_ds: gdal.Dataset = driver.Create(sla_path, ds_width, ds_height, bands=ds_band_num, eType=gdal.GDT_Float32)
    sla_ds.SetGeoTransform(ds_trans)
    sla_ds.SetProjection(ds_proj)
    sla_array = sla_ds.GetRasterBand(1).ReadAsArray()
    # get reference parameter
    ref_s_dict = {}
    for i in range(8):
        if i == 7:
            ref_s_dict[str(i + 1)] = ref_dictionary[str(11)]["5"]
        else:
            ref_s_dict[str(i + 1)] = ref_dictionary[str(i + 1)]["5"]
    # exchange and iteration(make sure all the index is right in your master optimising the speed)
    ref_ls = list(ref_s_dict.values())
    for i in tqdm(range(len(ref_ls))):
        if ref_ls[i] not in {1, 2, 3, 4, 5, 6, 7, 11}:
            if i == 7:
                lucc_array[lucc_array == 11] = ref_ls[7]
            else:
                lucc_array[lucc_array == i + 1] = ref_ls[i]
        else:
            for j in range(lucc_array.shape[1]):
                for k in range(lucc_array.shape[0]):
                    if lucc_array[k][j] == 11:
                        lucc_array[k][j] = ref_ls[i]
                    if lucc_array[k][j] == i + 1:
                        lucc_array[k][j] = ref_ls[i]
    lucc_array[lucc_array == 8] = 0
    sla_ds.WriteArray(lucc_array)
    sla_ds.FlushCache()
    del sla_ds
    print("SLA raster done")


def get_specific_lucc_dict():
    # get specific LUCC file list
    work_dir = r"F:\DATA\Vegetation_Resilience_D_DATA_C\0829_archive\LUCC(use)\LUCC_RESAMPLE"
    file_ls = workdir_filelist(work_dir)

    # construct ref dictionary
    ref_dict = {}
    json_path = r"/\VegetationResilience\REF_File\lut_0904_1.json"
    if os.path.exists(json_path):
        with open(json_path, 'r') as f:
            ref_dict = json.load(f)
    else:
        lut_path = r"C:\Users\PZH\Desktop\LUT_0904.xlsx"
        work_book = openpyxl.load_workbook(lut_path)
        work_sheet = work_book["Sheet1"]
        for i in range(8):
            para_dict = {}
            for j in range(12):
                para_dict[work_sheet.cell(j + 3, 2).value] = work_sheet.cell(j + 3, i + 3).value
            ref_dict[work_sheet.cell(2, i + 3).value] = para_dict
        ref_json = json.dumps(ref_dict)
        with open(json_path, 'w') as f:
            f.write(ref_json)
    return ref_dict


if __name__ == "__main__":
    # construct ref dictionary
    ref_dict = get_specific_lucc_dict()

    # get specific LUCC file list
    work_dir = r"D:\Data\VegetationResilienceDealing\Integrate_Output\LUCC(use)\LUCC_RESAMPLE"
    file_ls = workdir_filelist(work_dir)

    # read and creat the para-raster
    # parallel calculation
    out_dir = r"D:\Data\VegetationResilienceDealing\Integrate_Output\PARAMETER\SLA"
    para_ls = [(file_ls[i], out_dir, ref_dict) for i in range(len(file_ls))]
    with Pool(processes=8) as pool:
        results = pool.starmap(para_cal, para_ls)
