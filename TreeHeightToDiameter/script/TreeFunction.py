import math

import numpy as np
from openpyxl import load_workbook


class Trees:
    def __init__(self, height=None, diameter=None):
        self.height = height
        self.diameter = diameter

    @staticmethod
    def calculate_carbon(self, a, b, c, d, e):
        biomass_above = a * (self.diameter ** b)
        biomass_below = c * (self.diameter ** d)
        carbon = (biomass_above + biomass_below) * e
        return carbon


class SongLei(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height is not None:
            if self.height > 6.21:
                self.diameter = (self.height / 2.23902) ** (1 / 0.63384)
            else:
                self.diameter = (self.height / 1.58962) ** (1 / 0.84657)

    @staticmethod
    def calculate_area_carbon(F_array: np.ndarray, Slope_array: np.ndarray, H_array: np.ndarray):
        a = -2.0410
        b = 0.9597
        c = 1.2365
        S = (100 * F_array) / Slope_array
        carbon = (np.e ** a) * (S ** b) * (H_array ** c)
        return carbon


class LiLei(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height is not None:
            if self.height > 1.71:
                self.diameter = (self.height ** 0.5839 / 0.67313) ** (1 / 0.34574)
            else:
                self.diameter = (self.height ** 0.5839 / 0.91265) ** (1 / 0.35675)

    @staticmethod
    def calculate_area_carbon(F_array: np.ndarray, Slope_array: np.ndarray, H_array: np.ndarray):
        a = -2.0410
        b = 0.9597
        c = 1.2365
        S = (100 * F_array) / Slope_array
        carbon = (np.e ** a) * (S ** b) * (H_array ** c)
        return carbon


class ShanLei(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height is not None:
            if self.height > 4.82:
                self.diameter = ((self.height / 1.36904) ** (1 / 0.78216))
            else:
                self.diameter = ((self.height / 1.24783) ** (1 / 0.83975))

    @staticmethod
    def calculate_area_carbon(F_array: np.ndarray, Slope_array: np.ndarray, H_array: np.ndarray):
        a = -1.741
        b = 1.0419
        c = 1.2619
        S = (100 * F_array) / Slope_array
        carbon = (np.e ** a) * (S ** b) * (H_array ** c)
        return carbon


class YangShu(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height is not None:
            self.diameter = ((self.height + 0.0988) / 2.7536) ** (3 / 2)

    @staticmethod
    def calculate_area_carbon(F_array: np.ndarray, Slope_array: np.ndarray, H_array: np.ndarray):
        a = -1.0646
        b = 0.6774
        c = 1.4292
        S = (100 * F_array) / Slope_array
        carbon = (np.e ** a) * (S ** b) * (H_array ** c)
        return carbon


class ZhangShu(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height > 10.3:
            self.diameter = (self.height - 9.0409) / 0.1239
        else:
            self.diameter = math.e ** ((self.height - 6.228) / 1.757)

    @staticmethod
    def calculate_area_carbon(F_array: np.ndarray, Slope_array: np.ndarray, H_array: np.ndarray):
        a = -22.9503
        b = 5.8072
        c = 1.1774
        S = (100 * F_array) / Slope_array
        carbon = (np.e ** a) * (S ** b) * (H_array ** c)
        return carbon


class BaiLei(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height is not None:
            self.diameter = 1.377 * self.height + 2.307


class GuoShuLei(Trees):
    def __init__(self, height=None, diameter=None):
        super().__init__(height, diameter)

    def calculate_dia(self):
        if self.height is not None:
            self.diameter = (self.height - 0.738773) / 0.234835

    @staticmethod
    def calculate_area_carbon(F_array: np.ndarray, Slope_array: np.ndarray, H_array: np.ndarray):
        a = 17.2534
        b = -2.4699
        c = 0.5074
        S = (100 * F_array) / Slope_array
        carbon = (np.e ** a) * (S ** b) * (H_array ** c)
        return carbon


def find_carbon_coefficient(table: dict, number):
    # 导入一个Workbook
    coefficient = None
    for i in table.keys():
        try:
            if number == i:
                coefficient = table[i]
        except Exception as e:
            print(e)
    if coefficient is None:
        print("No matching species")
    else:
        return coefficient


def get_co_table(table_path: str):
    wb = load_workbook(filename=table_path)
    sheet = wb["江苏树种代码及计算参数"]
    coefficient = None
    co_dict = {}
    for i in range(1, sheet.max_row):
        co_dict[sheet.cell(i + 1, 1).value] = (sheet.cell(i + 1, 5).value, sheet.cell(i + 1, 6).value,
                                               sheet.cell(i + 1, 7).value, sheet.cell(i + 1, 8).value,
                                               sheet.cell(i + 1, 9).value)
    return co_dict
