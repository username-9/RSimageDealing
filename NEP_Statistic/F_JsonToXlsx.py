import json
import openpyxl

import pandas as pd

def json_to_xlsx(json_path, xlsx_path):
    json_file_path = json_path
    df = pd.read_json(json_path)

    # 将 DataFrame 写入 XLSX 文件
    xlsx_file_path = xlsx_path  # 替换为你希望输出的 XLSX 文件路径
    df.to_excel(xlsx_file_path, index=False)  # index=False 表示不将 DataFrame 的索引写入 XLSX 文件

    print(f"done\n{xlsx_file_path}")

def dict_to_xlsx(_dict, xlsx_path):
    df = pd.DataFrame(_dict)
    df.to_excel(xlsx_path, index=False)
    print(f"done\n{xlsx_path}")


if __name__ == '__main__':
    json_path = r"C:\Users\PZH\Desktop\ProgrammeGuangdong\1029\Xian_Carbon_STATISTIC.json"
    xlsx_path = r"C:\Users\PZH\Desktop\ProgrammeGuangdong\1029\Xian_Carbon_STATISTIC.xlsx"
    # json_to_xlsx(json_path, xlsx_path)
    data_dict = json.load(open(json_path, encoding='utf-8'))
    contain_dict = {}
    area_dict = {}

    for key, v_dict in data_dict.items():
        temp_dict_1 = {}
        temp_dict_2 = {}
        for k, v in v_dict.items():
            temp_dict_1[k] = v[0]
            temp_dict_2[k] = v[1]
        contain_dict[key] = temp_dict_1
        area_dict[key] = temp_dict_2

    # write in XLSX
    dict_to_xlsx(contain_dict, xlsx_path)
    temp_path = r"C:\Users\PZH\Desktop\ProgrammeGuangdong\1029\Xian_Area.xlsx"
    dict_to_xlsx(area_dict, temp_path)
    row_name = list(data_dict["2000"].keys())
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active
    max_con = worksheet.max_column
    for i in range(len(row_name)):
        worksheet.cell(i+2, max_con+1).value = row_name[i]
    workbook.save(xlsx_path)


