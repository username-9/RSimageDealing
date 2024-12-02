import csv
import typing

import openpyxl
import os


def load_xlsx_for_x_y_data(file_path, sheet_name, x_field, y_field, value_field: list or tuple) -> typing.Generator:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    if isinstance(value_field, tuple) or isinstance(value_field, list):
        value_field = [value_field]

    headers = {cell.value: cell.column for cell in sheet[1]}

    if x_field not in headers or y_field not in headers or value_field not in headers:
        raise ValueError("Invalid value for x_field, y_field, value_field")

    x_col_idx = headers[x_field]
    y_col_idx = headers[y_field]
    value_col_idx = [headers[value_field[i]] for i in range(len(value_field))]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        try:
            x_value = row[x_col_idx - 1]
            y_value = row[y_col_idx - 1]
            attribute_value = (row[v_col_idx - 1] for v_col_idx in value_col_idx)
            yield x_value, y_value, attribute_value
        except ValueError as e:
            print(f"Skipping row due to invalid value or index error: {e}")


def load_csv_for_x_y_data(file_path,x_field, y_field, value_field) -> typing.Generator:
    with open(file_path, newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        # x_index = reader.fieldnames.index(x_field)
        # y_index = reader.fieldnames.index(y_field)
        # value_index = (reader.fieldnames.index(value_field[_i]) for _i in range(len(value_field)))
        if not (isinstance(value_field, tuple) or isinstance(value_field, list)):
            value_field = [value_field]

        for row in reader:
            try:
                x_value = eval(row[x_field])
                y_value = eval(row[y_field])
                attribute_value = [eval(row[index]) for index in value_field]
                yield x_value, y_value, attribute_value
            except ValueError as e:
                print(f"Skipping row due to invalid value or index error: {e}")


if __name__ == "__main__":
    # # TEST load_xlsx_for_x_y_data 24 10 31
    # generate_test_tuple = load_csv_for_x_y_data(r"C:\Users\PZH\Desktop\渔网与熵\point_cloud_count_percent_entropy_new.csv",
    #                                              "mean_x",
    #                                              "mean_y",
    #                                              "Entropy")
    # for i in generate_test_tuple:
    #     print(i)
    pass