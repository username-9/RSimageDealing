import re
import typing

import numpy as np
import openpyxl
import tqdm
from osgeo import gdal, ogr

from UtilitiesForProcessingImage.BasicUtility.UtilityFunction import workdir_filelist


def raster_read(file_path: str, deal_type: str = gdal.GA_ReadOnly) -> gdal.Dataset:
    """
    read raster with more information ,also can use the origin method of gdal (gdal.Open())
    :param file_path: the raster file path
    :param deal_type: the type of how to read dataset(find in gdal manual)
    :return: gdal.Dataset
    """
    try:
        src = gdal.Open(file_path, deal_type)
        # read prac.
        rows = src.RasterYSize
        cols = src.RasterXSize
        bands = src.RasterCount
        print(f"[{file_path}]\nrows={rows}  column={cols}  bands={bands}")

        # get geo transform info.
        geo_info = src.GetGeoTransform()
        print(f"Geograph information(image to projection): {geo_info}")

        # get geo projection info.
        geo_proj = src.GetProjection()
        print(f"Projection information(projection to geo_coordination): {geo_proj}")

        # # get band information
        # fir_band = src.GetRasterBand(1)
        # the_max = fir_band.GetMaximum()
        # the_min = fir_band.GetMinimum()
        # print(fir_band, the_max, the_min)

        # read as array
        # fir_array = fir_band.ReadAsArray()
        # print(fir_array.shape)

        # show
        # mean = np.mean(fir_array)
        # std = np.std(fir_array)
        # plt.imshow(fir_array,
        #            vmin=mean - std * 2,
        #            vmax=mean + std * 2,
        #            cmap="gray")
        # plt.show()

        # mult_band show
        # ……
        return src
    except Exception as e:
        print(f"{e}\nmay be the driver cannot be used")


def vector_read(file_path: str, operate_type: int = 0, file_type: str = "ESRI Shapefile") -> gdal.ogr.DataSource:
    """
    read vector data
    :param operate_type: 
    :param file_path: vector file path
    :param file_type: type of vector file(default is ESRI Shapefile)
    :return: gdal.Dataset
    """
    gdal.SetConfigOption("GDAL_FILENAME_IS_UTF8", "YES")
    gdal.SetConfigOption("SHAPE_ENCODING", "UTF-8")
    ogr.RegisterAll()
    driver = ogr.GetDriverByName(file_type)
    try:
        src = driver.Open(file_path, operate_type)
        return src
    except Exception as e:
        print(f"{e}\nmay be the driver cannot be used")


def hdf_read(file_path: str, deal_type: str = gdal.GA_ReadOnly) -> gdal.Dataset or list:
    """
    read HDF file
    :param file_path: the source file path
    :param deal_type: type of gdal.Open
    :return: gdal.Dataset
    """
    try:
        ds = gdal.Open(file_path, deal_type)
        sub_datasets = ds.GetSubDatasets()
        if not sub_datasets:
            return ds
        else:
            print('Number of sub-datasets: {}'.format(len(sub_datasets)))
            for sd in sub_datasets:
                print('Name: {0}\nDescription:{1}\n'.format(*sd))
            return sub_datasets
    except Exception as e:
        print(f"{e}\nSome wrongs happened in reading hdf")


def read_band_scale_offset(ds: gdal.Dataset, bands: list = None):
    """
    read scale and offset information of dataset
    :param ds: gdal dataset
    :param bands: bands which need to be read
    :return: None
    """
    scale = []
    offset = []
    if bands is None:
        try:
            for i in range(ds.RasterCount):
                ds_band = ds.GetRasterBand(i + 1)
                scale_1 = ds_band.GetScale()
                offset_1 = ds_band.GetOffset()
                scale.append(scale_1)
                offset.append(offset_1)
                del ds_band
            return scale, offset
        except Exception as e:
            print(f"{e}\ngetting all band information is wrong")
    else:
        try:
            for i in bands:
                ds_band = ds.GetRasterBand(i)
                scale_1 = ds_band.GetScale()
                offset_1 = ds_band.GetOffset()
                scale.append(scale_1)
                offset.append(offset_1)
                del ds_band
            return scale, offset
        except Exception as e:
            print(f"{e}\nSome wrongs happened in reading band information")


def get_raster_nodata(ds: gdal.Dataset, bands: list = None):
    nodata = []
    if bands is None:
        try:
            for i in range(ds.RasterCount):
                ds_band: gdal.Band = ds.GetRasterBand(i + 1)
                nodata = ds_band.GetNoDataValue()
                del ds_band
            return nodata
        except Exception as e:
            print(f"{e}\ngetting all band nodata information false")
    else:
        try:
            for i in bands:
                ds_band = ds.GetRasterBand(i)
                nodata = ds_band.GetNoDataValue()
                del ds_band
            return nodata
        except Exception as e:
            print(f"{e}\nSome wrongs happened in getting defined bands information")


def from_geo_coordination_to_picture_coordination(longitude: float or str, latitude: float or str, transform: str,
                                    degree_minute_second = False, bands: list = None) -> tuple:
    """
    get value with geograph coordination
    :param transform: geographic transformation
    :param longitude: longitude
    :param latitude: latitude
    :param degree_minute_second: whether using str format as "degree-minute-second"
    :param bands: which bands need to be read
    :return: the result value getting from raster
    """

    if degree_minute_second:
        re_longitude = re.split(r'[°′″]+', longitude)
        lon_degrees = float(re_longitude[0])
        lon_minutes = float(re_longitude[1])
        lon_seconds = float(re_longitude[2])
        sign = 1
        if 'W' in longitude:
            sign = -1
        else:
            sign = 1
        longitude = lon_degrees + sign * (lon_minutes / 60 + lon_seconds / 3600)
        re_latitude =  re.split(r'[°′″]+', latitude)
        lat_degrees = float(re_latitude[0])
        lat_minutes = float(re_latitude[1])
        lat_seconds = float(re_latitude[2])
        if 'S' in latitude:
            sign = -1
        else:
            sign = 1
        latitude = lat_degrees + sign * (lat_minutes / 60 + lat_seconds / 3600)
    x_offset, y_offset = gdal.ApplyGeoTransform(transform, longitude, latitude)
    return round(x_offset), round(y_offset)


def get_value_from_geo_coordination(src: gdal.Dataset, longitude: float or str, latitude: float or str,
                                    transform: str, degree_minute_second = False, bands: list = None) -> typing.Any:
    re_value = []
    x, y = from_geo_coordination_to_picture_coordination(longitude, latitude, transform, degree_minute_second, bands)
    if bands is None:
        bands = [1]
    for i in bands:
        arr = src.GetRasterBand(i).ReadAsArray()
        value = arr[y, x]
        re_value.append(value)
    if len(re_value) == 0:
        re_value = re_value[0]
    return re_value


def raster_to_array(file_path: str) -> np.ndarray:
    """
    get array from raster (file path)
    :param file_path: raster file path
    :return: the array contained by raster
    """
    _ds = gdal.Open(file_path)
    array = _ds.ReadAsArray()
    del _ds
    return array


if __name__ == '__main__':
    # _src = raster_read(r"C:\Users\Administrator\Downloads\MYDLT1F.20140601.CN.LTN.MAX.V2.TIF")
    # del _src
    work_dir = r"F:\DATA\Vegetation_Resilience_D_DATA_C\NPP_MONTHLY"
    file_list = workdir_filelist(work_dir)
    value_dict = {}
    lon ="116°34′12.72″E"
    lat = "36°49′44.4″N"
    for file in tqdm.tqdm(file_list):
        ds: gdal.Dataset = gdal.Open(file)
        tran = ds.GetGeoTransform()
        value = get_value_from_geo_coordination(ds, lon, lat, tran, degree_minute_second=True)
        year_month = file[4:12]
        value_dict[year_month] = value
    xlsx = r"C:\Users\PZH\Desktop\verify_statioin.xlsx"
    work_book = openpyxl.load_workbook(xlsx)
    ws = work_book["Sheet1"]
    for index, item in tqdm.tqdm(enumerate(value_dict.items())):
        y= item[0]
        v = item[1][0]
        ws.cell(row=index+1, column=1).value = y
        ws.cell(row=index+1, column=2).value = v
    work_book.save(xlsx)
    print("done")

